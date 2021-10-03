#! python
# -*- coding:utf-8 -*-

import os
import re
import sys
import json
import openpyxl

from slpp.slpp import slpp as lua

TYPES = {"num": True, "str": True, "json": True, "lua": True}

# 用于判断是否整数的正则
INT_RE = re.compile(r"^[-]?\d+$")


def is_int_str(s):
    return INT_RE.match(s) is not None

# 如果不是字符串，则转换为字符串
def to_unicode_str(val):
    if isinstance(val, str):
        return val
    else:
        return str(val)

# 根据配置类型，把值转换为对应的数据类型
def to_type_value(val_type, val):
    # openpyxl 返回的value有多种类型，需要我们自己转换
    if None == val:
        return None
    elif "num" == val_type:
        if isinstance(val, int) or isinstance(val, float):
            return val

        if is_int_str(val):
            return int(val)

        return float(val)
    elif "str" == val_type:
        return to_unicode_str(val)
    elif "json" == val_type:
        # 空的json结构不导出，避免占用不必要的内存
        if "[]" == val or "{}" == val:
            return None

        return json.loads(val)
    elif "lua" == val_type:
        # 空的lua结构不导出，避免占用不必要的内存
        if "{}" == val:
            return None

        return lua.decode(val)
    else:
        raise_error("invalid type", val_type)

# 导出字段的名字、参数等
class Field(object):

    def __init__(self, t, o, n):
        self.name = n
        self.opt  = o
        self.type = t

# 继承object类，以解决在python2中的错误：TypeError: must be type, not classobj


class Sheet(object):

    def __init__(self, base_name, wb_sheet):
        self.fields = []  # 各列字段名

        self.wb_sheet = wb_sheet
        self.base_name = base_name

        self.index = -1  # 索引数量，-1表示kv模式
        self.dir_path = None  # 导出目录
        self.file_path = None  # 导出文件名

        # 记录出错时的行列，方便定位问题
        self.error_row = 0
        self.error_col = 0
        self.error_val = None

    # 记录出错位置
    def mark_error_pos(self, row, col):
        if row > 0:
            self.error_row = row
        if col > 0:
            self.error_col = col

    # 发起一个解析错误
    def raise_error(self, what, val):
        excel_info = "DOC:{0},SHEET:{1},ROW:{2},COLUMN:{3}, VAL: {4}".format(
            self.base_name, self.wb_sheet.title,
            self.error_row, self.error_col, self.error_val)
        raise Exception(what, val, excel_info)

    def to_value(self, val_type, val):
        try:
            self.error_val = val
            return to_type_value(val_type, val)
        except Exception:
            t, e = sys.exc_info()[:2]
            self.raise_error("ConverError", e)


    # 解析各字段的类型
    def to_field_type(self, row, col):
        self.mark_error_pos(row, col)
        value = self.wb_sheet.cell(row=row, column=col).value

        # 如果未指定类型，则表示后面的数据都不导出了
        if value == None:
            return None

        if not TYPES.get(value):
            self.raise_error("invalid type", value)

        return value

    # 解析各字段的类型
    def to_field_name(self, row, col):
        self.mark_error_pos(row, col)
        value = self.wb_sheet.cell(row=row, column=col).value
        if value == None:
            return None

        return to_unicode_str(value)

    # 导出选项：是导出服务端还是客户端
    def to_field_opt(self, row, col):
        self.mark_error_pos(row, col)
        value = self.wb_sheet.cell(row=row, column=col).value

        # 中间可能有些注释字段，不需要导出
        if value == None:
            return 0
        # 服务器占第一位，客户端第二位
        elif "s" == value:
            return 1
        elif "c" == value:
            return 2
        elif "sc" == value or "cs" == value:
            return 3
        else:
            self.raise_error("invalid option", value)

    # 导出单个格子中的数据
    def to_field_value(self, srv_ctx, clt_ctx, field, row, col):
        # 不需要导出，可能是注释
        opt = field.opt
        if 0 == opt: return None

        self.mark_error_pos(row, col)
        value = self.wb_sheet.cell(row=row, column=col).value

        # 可能返回None，可能是没配置或者空lua表之类的
        value = self.to_value(field.type, value)
        if None == value: return None

        name = field.name
        if opt & 0x1: srv_ctx[name] = value
        if opt & 0x2: clt_ctx[name] = value

        return value

    # 解析基础信息，如目录、导出文件名等
    def decode_info(self):
        row = 1
        col = 1
        self.index = self.wb_sheet.cell(
            row = row, column = col).value
        self.dir_path = self.wb_sheet.cell(
            row = row, column = col + 1).value
        self.file_path = self.wb_sheet.cell(
            row = row, column = col + 2).value

    # 解析一个表格
    def decode_sheet(self):
        wb_sheet = self.wb_sheet

        self.decode_info()  # 解析基础信息，如目录、导出文件名等
        self.decode_field()  # 解析字段类型、导出参数，是否导出服务端、客户端
        self.decode_ctx()  # 解析内容

        print("    decode sheet %s done" % wb_sheet.title.ljust(24, "."))

# 导出数组类型配置


class ArraySheet(Sheet):

    def __init__(self, base_name, wb_sheet):
        # 记录导出各行的内容
        self.srv_ctx = None
        self.clt_ctx = None

        super(ArraySheet, self).__init__(base_name, wb_sheet)

    # 解析一个字段信息，名字、类型、导出参数等
    def decode_field(self):
        row = 3
        beg_col = 1

        has = set()
        for col in range(beg_col, self.wb_sheet.max_column + 1):
            f_type = self.to_field_type(row, col)
            if not f_type:
                break

            f_opt = self.to_field_opt(row + 1, col)
            f_name = self.to_field_name(row + 2, col)
            if f_name in has:
                self.raise_error("field name dumplicate", f_name)

            self.fields.append(Field(f_type, f_opt, f_name))

    # 解析出一行的内容
    def decode_row(self, row):
        srv_row = {}
        clt_row = {}

        # 第一列没数据，从第二列开始解析
        beg_col = 1
        for col in range(beg_col, len(self.fields) + beg_col):
            field = self.fields[col - 1]
            self.to_field_value(srv_row, clt_row, field, row, col)

        return srv_row, clt_row  # 返回一个tuple

    # 生成第N层索引数据
    def make_index(self, ctx, i, max_i):
        name = self.fields[i].name

        index_ctx = {}
        for v in ctx:
            val = v.get(name)
            if not val:
                self.raise_error("index no value set", name)

            # 如果使用了索引，则索引得到的内容必须是唯一的
            old = index_ctx.get(val)
            if i >= max_i:
                if old:
                    self.raise_error(
                        "dumplicate index value", name + ": " + str(val))
                else:
                    index_ctx[val] = v
            else:
                # 还不是最后一层索引，存到一个数组等后续处理
                if not old:
                    old = []
                    index_ctx[val] = old
                old.append(v)

        if i < max_i:
            new_ctx = {}
            for k, v in index_ctx.items():
                new_ctx[k] = self.make_index(v, i + 1, max_i)

            return new_ctx
        else:
            return index_ctx

    # 解析导出的内容
    def decode_ctx(self):
        srv_ctx = []
        clt_ctx = []
        # 前面几行分别是：基础信息、注释、类型、选项、字段名
        beg_row = 6
        for row in range(beg_row, self.wb_sheet.max_row + 1):
            srv_row, clt_row = self.decode_row(row)

            has = False
            # 不为空才追加
            if any(srv_row):
                has = True
                srv_ctx.append(srv_row)
            if any(clt_row):
                has = True
                clt_ctx.append(clt_row)

            # 遇到空行后面的不再导出
            if not has: break

        self.error_val = None
        if self.index > 0:
            if self.index > len(self.fields):
                self.raise_error("index lager than field count", self.index)

            if len(srv_ctx) > 0:
                self.srv_ctx = self.make_index(srv_ctx, 0, self.index - 1)
            if len(clt_ctx) > 0:
                self.clt_ctx = self.make_index(clt_ctx, 0, self.index - 1)
        else:
            self.srv_ctx = srv_ctx
            self.clt_ctx = clt_ctx

# 导出object类型的结构


class ObjectSheet(Sheet):

    def __init__(self, base_name, wb_sheet):
        # 记录导出各行的内容
        self.srv_ctx = {}
        self.clt_ctx = {}

        super(ObjectSheet, self).__init__(base_name, wb_sheet)

    # 解析一个字段信息，名字、类型、导出参数等
    def decode_field(self):
        col = 2  # 类型所在列
        beg_row = 2

        has = set()
        for row in range(beg_row, self.wb_sheet.max_row + 1):
            f_type = self.to_field_type(row, col)
            if not f_type:
                break

            f_opt = self.to_field_opt(row, col + 1)
            f_name = self.to_field_name(row, col + 2)
            if f_name in has:
                self.raise_error("field name dumplicate", f_name)

            self.fields.append(Field(f_type, f_opt, f_name))

    # 解析表格的所有内容
    def decode_ctx(self):
        col = 5 # 数据所在列
        beg_row = 2
        for row in range(beg_row, len(self.fields) + beg_row):
            # 第一行没数据，-1.数组下标从0开始，但excel坐标从1开始，再-1
            field = self.fields[row - 2]
            
            # 不需要导出，可能是注释
            opt = field.opt
            if 0 == opt:
                continue

            self.mark_error_pos(row, col)
            value = self.wb_sheet.cell(row=row, column=col).value

            value = self.to_value(field.type, value)
            # 空值，可能是没配置或者空lua表之类的
            if None == value:
                continue

            name = field.name
            if opt & 0x1: self.srv_ctx[name] = value
            if opt & 0x2: self.clt_ctx[name] = value


class ExcelDoc:

    def __init__(self, file, path):
        self.file = file
        self.path = path

    # 是否需要解析
    def need_decode(self, wb_sheet):
        sheet_val = wb_sheet.cell(row = 1, column = 1).value

        if not isinstance(sheet_val, int):
            return -2

        # 判断索引是否正确，-1表示kv模式，其他表示索引数量
        if -1 == sheet_val:
            return -1
        elif sheet_val >= 0:
            if sheet_val > 8:
                print("sheet %s too many index, abort" %
                      wb_sheet.title.ljust(24, "."))
                return -2
            return sheet_val
        else:
            return -2

    def decode(self, srv_path, clt_path, SrvWriter, CltWriter):
        print("start decode %s ..." % self.file)

        # base_name = os.path.splitext(self.file)[0]  # 去除后缀
        wb = openpyxl.load_workbook(self.path)

        for wb_sheet in wb.worksheets:
            index = self.need_decode(wb_sheet)
            if index < -1:
                print("    sheet %s no need to decode,abort" %
                      wb_sheet.title.ljust(24, "."))
                continue
            elif -1 == index:
                sheet = ObjectSheet(self.file, wb_sheet)
            else:
                sheet = ArraySheet(self.file, wb_sheet)

            sheet.decode_sheet()
            if SrvWriter:
                writer = SrvWriter(self.file, wb_sheet.title)
                writer.write_to_file(sheet.srv_ctx,
                    srv_path, sheet.dir_path, sheet.file_path)

            if CltWriter:
                writer = CltWriter(self.file, wb_sheet.title)
                writer.write_to_file(sheet.clt_ctx,
                    clt_path, sheet.dir_path, sheet.file_path)

